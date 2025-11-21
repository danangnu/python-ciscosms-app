#!/usr/bin/env python3
import re, time, socket
from pathlib import Path
from datetime import datetime
import pandas as pd
import paramiko

# ====== CONFIG ======
HOST = "192.168.206.1"
USERNAME = "admin"
PASSWORD = "Bryan2011"
SSH_TIMEOUT = 12
CMD_TIMEOUT = 25

# Output locations
OUT_DIR = Path.cwd()
XLSX_PATH = (OUT_DIR / "router_versions.xlsx").resolve()
CSV_PATH  = (OUT_DIR / "router_versions.csv").resolve()

# Base columns
BASE_COLS = ["host","model_image","version","raw_first_line","status"]

# LTE columns we’ll try to populate if Cellular exists
LTE_COLS = [
    "lte_present","lte_slot","lte_vendor","lte_model","lte_hw",
    "lte_fw","lte_radio","lte_carrier","lte_pri","lte_imei","lte_imsi","lte_iccid"
]

COLUMNS = BASE_COLS + LTE_COLS

# ====== Regex ======
IMAGE_RE   = re.compile(r"\(([^)]+UNIVERSALK9[^)]*)\)", re.IGNORECASE)
VERSION_RE = re.compile(r"Version\s+([0-9A-Za-z().\-+]+)", re.IGNORECASE)
IOSXE_VER  = re.compile(r"Cisco IOS XE Software.*Version\s+([0-9A-Za-z().\-+]+)", re.IGNORECASE)

# Cellular presence / slot guessers
CELL_LINE  = re.compile(r"^Cellular(?P<slot>[\d/]+)\s", re.IGNORECASE)
IMEI_RE    = re.compile(r"\bIMEI[\s:]+([0-9]{8,20})", re.IGNORECASE)
IMSI_RE    = re.compile(r"\bIMSI[\s:]+([0-9]{8,20})", re.IGNORECASE)
ICCID_RE   = re.compile(r"\bICCID[\s:]+([0-9A-Za-z]+)", re.IGNORECASE)
FW_RE      = re.compile(r"(?:Firmware|Revision|SW Version|Software Version)\s*[:=]\s*([^\r\n]+)", re.IGNORECASE)
HW_RE      = re.compile(r"(?:Hardware|HW Version)\s*[:=]\s*([^\r\n]+)", re.IGNORECASE)
RADIO_RE   = re.compile(r"(?:Radio|Technology|Access Tech|Current System Time|Sys Mode)\s*[:=]\s*([^\r\n]+)", re.IGNORECASE)
CARRIER_RE = re.compile(r"(?:Operator|Network Operator|PLMN|Provider)\s*[:=]\s*([^\r\n]+)", re.IGNORECASE)
PRI_RE     = re.compile(r"(?:Carrier Config|PRI|Carrier PRI|Config Version)\s*[:=]\s*([^\r\n]+)", re.IGNORECASE)

# Common vendor/model clues in various outputs
VENDOR_MODEL_HINTS = [
    re.compile(r"(Sierra Wireless|Qualcomm|Huawei|Telit|Quectel|Novatel|Ericsson|Gobi|Gemalto)", re.IGNORECASE),
    re.compile(r"\b(EM7\d{3}|MC7\d{3}|MC77\d{2}|EC2\d|EC25|EC20|EG25|EM7455|EM7430|MC7350|MC7700|MC7710|MC7304|ME909|MU7\d{2}|LE9\d{2}|LN9\d{2})\b", re.IGNORECASE),
]

# ====== SSH helpers ======
def ssh_collect(commands):
    """
    Open a single shell, send commands, return combined buffer.
    """
    cli = paramiko.SSHClient()
    cli.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    cli.connect(
        HOST, username=USERNAME, password=PASSWORD,
        look_for_keys=False, allow_agent=False,
        timeout=SSH_TIMEOUT, auth_timeout=SSH_TIMEOUT
    )

    ch = cli.invoke_shell()
    ch.settimeout(2.0)

    def send(line: str):
        ch.send(line + "\n")
        time.sleep(0.15)

    send("")  # get prompt
    send("terminal length 0")
    for c in commands:
        send(c)

    buf, start = "", time.time()
    while time.time() - start < CMD_TIMEOUT:
        try:
            chunk = ch.recv(65535).decode("utf-8", errors="ignore")
            if chunk:
                buf += chunk
                if re.search(r"\n[\w\-\.\:]+[#>]\s*$", buf):
                    time.sleep(0.15)
                    break
            else:
                time.sleep(0.1)
        except socket.timeout:
            pass

    ch.close(); cli.close()
    return buf

# ====== Parsers ======
def parse_ios_version(buffer: str):
    first = ""
    for line in buffer.splitlines():
        if "Cisco IOS" in line or "Cisco IOS XE" in line or "Cisco IOS Software" in line:
            first = line.strip()
            break

    model_image = ""
    m_img = IMAGE_RE.search(buffer)
    if m_img:
        model_image = m_img.group(1).strip()

    version = ""
    m_ver = VERSION_RE.search(first or buffer)
    if m_ver:
        version = m_ver.group(1).strip()
    else:
        m = IOSXE_VER.search(buffer)
        if m:
            version = m.group(1).strip()

    return first, model_image, version

def detect_cellular_presence(ipbr: str):
    """
    From 'show ip interface brief', find any Cellular lines and slot strings.
    Returns (present: bool, first_slot: str|None)
    """
    slots = []
    for line in ipbr.splitlines():
        m = CELL_LINE.search(line.strip())
        if m:
            slots.append(m.group("slot"))
    return (len(slots) > 0, (slots[0] if slots else None))

def parse_lte_info(text: str, slot_guess: str | None):
    """
    Best-effort extraction of LTE details from any of the cellular/ctrl/inventory outputs.
    """
    lte = {
        "lte_present": "no",
        "lte_slot": slot_guess or "",
        "lte_vendor": "",
        "lte_model": "",
        "lte_hw": "",
        "lte_fw": "",
        "lte_radio": "",
        "lte_carrier": "",
        "lte_pri": "",
        "lte_imei": "",
        "lte_imsi": "",
        "lte_iccid": "",
    }

    # any hit in the cellular outputs implies presence
    if any(k in text.lower() for k in ["cellular", "modem", "imei", "imsi", "iccid"]):
        lte["lte_present"] = "yes"

    # identifiers
    m = IMEI_RE.search(text);   lte["lte_imei"]   = m.group(1) if m else ""
    m = IMSI_RE.search(text);   lte["lte_imsi"]   = m.group(1) if m else ""
    m = ICCID_RE.search(text);  lte["lte_iccid"]  = m.group(1) if m else ""

    m = FW_RE.search(text);     lte["lte_fw"]     = (m.group(1).strip() if m else "")
    m = HW_RE.search(text);     lte["lte_hw"]     = (m.group(1).strip() if m else "")
    m = RADIO_RE.search(text);  lte["lte_radio"]  = (m.group(1).strip() if m else "")
    m = CARRIER_RE.search(text);lte["lte_carrier"]= (m.group(1).strip() if m else "")
    m = PRI_RE.search(text);    lte["lte_pri"]    = (m.group(1).strip() if m else "")

    # vendor & model hints anywhere
    vendor = ""
    model  = ""
    for rx in VENDOR_MODEL_HINTS:
        for hit in rx.finditer(text):
            val = hit.group(0).strip()
            # naive split of "Sierra Wireless MC7710" kinds of lines
            if not vendor and any(v in val.lower() for v in ["sierra","qualcomm","huawei","telit","quectel","novatel","ericsson","gobi","gemalto"]):
                vendor = val if " " not in vendor else vendor
            if not model and re.search(r"(EM7\d{3}|MC7\d{3}|MC77\d{2}|EC2\d|EC25|EC20|EG25|EM7455|EM7430|MC7350|MC7700|MC7710|MC7304|ME909|MU7\d{2}|LE9\d{2}|LN9\d{2})", val, re.IGNORECASE):
                model = val
    lte["lte_vendor"] = vendor
    lte["lte_model"]  = model

    # If inventory mentions manufacturer/product name
    inv_vendor = re.search(r"VID\s*[:=]\s*([^\r\n]+)", text, re.IGNORECASE)
    inv_name   = re.search(r"(?:NAME|DESC)\s*[:=]\s*\"?([^\r\n\"]+)\"?", text, re.IGNORECASE)
    if inv_vendor and not lte["lte_vendor"]:
        lte["lte_vendor"] = inv_vendor.group(1).strip()
    if inv_name and not lte["lte_model"]:
        name = inv_name.group(1).strip()
        # if name carries both, try to split
        if not lte["lte_vendor"] and any(k in name.lower() for k in ["sierra","qualcomm","huawei","telit","quectel","novatel","ericsson","gemalto"]):
            lte["lte_vendor"] = name
        else:
            lte["lte_model"] = name

    return lte

# ====== Append-safe writers ======
def append_csv(path: Path, row: dict):
    header = not path.exists()
    df = pd.DataFrame([row], columns=COLUMNS)
    with open(path, "a", encoding="utf-8", newline="") as f:
        df.to_csv(f, header=header, index=False)

def append_xlsx(path: Path, row: dict):
    try:
        import openpyxl
        if path.exists():
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            # Write header only if first cell is empty and sheet is blank
            if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
                ws.append(COLUMNS)
            ws.append([row.get(c, "") for c in COLUMNS])
            wb.save(path)
        else:
            pd.DataFrame([row], columns=COLUMNS).to_excel(path, index=False)
    except ImportError:
        # Fallback: read + concat + rewrite (still preserves history)
        if path.exists():
            old = pd.read_excel(path)
            new = pd.concat([old, pd.DataFrame([row], columns=COLUMNS)], ignore_index=True)
        else:
            new = pd.DataFrame([row], columns=COLUMNS)
        new.to_excel(path, index=False)

# ====== Main flow ======
def main():
    # 1) Get version info and interface list first
    combined = ssh_collect(["show version", "show ip interface brief"])
    first_line, model_image, version = parse_ios_version(combined)

    # 2) Cellular presence?
    present, slot_guess = detect_cellular_presence(combined)

    lte_info = {k: "" for k in LTE_COLS}
    lte_info["lte_present"] = "yes" if present else "no"
    lte_info["lte_slot"] = slot_guess or ""

    # 3) If present, probe likely commands to extract make/model/details
    if present:
        probe_cmds = []
        # try specific slot forms first if we can guess
        if slot_guess:
            probe_cmds += [f"show cellular {slot_guess} all", f"show controllers cellular {slot_guess}"]
        # also try common defaults
        probe_cmds += [
            "show cellular 0 all",
            "show cellular 0/0/0 all",
            "show controllers cellular 0/0/0",
            "show inventory",
        ]
        lte_buf = ssh_collect(probe_cmds)
        parsed = parse_lte_info(lte_buf, slot_guess)
        lte_info.update(parsed)

    row = {
        "host": HOST,
        "model_image": model_image,
        "version": version,
        "raw_first_line": first_line,
        "status": "ok" if (model_image or version) else "parse_failed",
        **lte_info,
    }

    print(pd.DataFrame([row], columns=COLUMNS).to_string(index=False))
    append_csv(CSV_PATH, row)
    append_xlsx(XLSX_PATH, row)
    print(f"\nAppended to:\n  {CSV_PATH}\n  {XLSX_PATH}")

if __name__ == "__main__":
    main()